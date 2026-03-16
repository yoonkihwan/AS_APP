from django.contrib import admin
from django.urls import path, reverse
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponseRedirect
from unfold.admin import ModelAdmin, TabularInline
from unfold.sites import UnfoldAdminSite
from unfold.decorators import action as unfold_action, display
from django.utils import timezone
from django.utils.html import format_html
from .models import Inventory, InventoryBatch, InboundInventory, OutboundInventory
from .forms import InventoryForm
from master_data.models import OutsourceCompany, Company, Tool, Brand

class ToolInventoryAdminSite(UnfoldAdminSite):
    site_title = "장비/툴 관리 시스템"
    site_header = "장비/툴 관리 대시보드"
    site_url = None
    index_title = "대시보드"
    index_template = "tool_inventory/index.html"

    def app_index(self, request, app_label, extra_context=None):
        return HttpResponseRedirect(reverse('tool_admin:index'))

    def get_urls(self):
        custom_urls = [
            path('dashboard/stock-pdf/', self.admin_view(self.dashboard_stock_pdf), name='dashboard_stock_pdf'),
            path('dashboard/history-pdf/', self.admin_view(self.dashboard_history_pdf), name='dashboard_history_pdf'),
        ]
        return custom_urls + super().get_urls()

    def dashboard_stock_pdf(self, request):
        """대시보드 원클릭: 현재 재고 현황 PDF 다운로드"""
        import io, os
        from django.http import FileResponse
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from django.db.models import Q
        from master_data.models import Tool
        from .models import Inventory
        from django.utils import timezone

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=40)
        elements = []

        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'
        try:
            font_path = "c:\\Windows\\Fonts\\malgun.ttf"
            font_path_bd = "c:\\Windows\\Fonts\\malgunbd.ttf"
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('Malgun', font_path))
                font_name = 'Malgun'
                font_name_bold = 'Malgun' # Fallback
            if os.path.exists(font_path_bd):
                pdfmetrics.registerFont(TTFont('MalgunBold', font_path_bd))
                font_name_bold = 'MalgunBold'
        except Exception:
            pass

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Normal'],
            fontName=font_name_bold, fontSize=18, alignment=1, spaceAfter=20
        )
        date_style = ParagraphStyle(
            'CustomDate', parent=styles['Normal'],
            fontName=font_name, fontSize=10, textColor=colors.gray, alignment=2, spaceAfter=15
        )
        normal_style = ParagraphStyle(
            'CustomNormal', parent=styles['Normal'],
            fontName=font_name, fontSize=9, leading=14
        )
        bold_style = ParagraphStyle(
            'CustomBold', parent=styles['Normal'],
            fontName=font_name_bold, fontSize=10, leading=14
        )
        center_style = ParagraphStyle(
            'CustomCenter', parent=normal_style, alignment=1
        )

        elements.append(Paragraph("현재 재고 현황", title_style))
        elements.append(Paragraph(f"출력일시: {timezone.now().strftime('%Y-%m-%d %H:%M')}", date_style))

        tools_with_stock = Tool.objects.filter(
            inventory__status='재고'
        ).distinct().select_related('brand').order_by('brand__name', 'model_name')

        data = []
        data.append([
            Paragraph("<b>품목명 (Model)</b>", ParagraphStyle('H1', parent=normal_style, fontName=font_name_bold)),
            Paragraph("<b>수량 (Qty)</b>", ParagraphStyle('H2', parent=center_style, fontName=font_name_bold)),
            Paragraph("<b>시리얼 내역 (S/N)</b>", ParagraphStyle('H3', parent=normal_style, fontName=font_name_bold))
        ])

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#94a3b8')),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
        ])

        current_brand = None
        total_qty = 0
        row_idx = 1

        for tool in tools_with_stock:
            if tool.brand != current_brand:
                current_brand = tool.brand
                brand_name = current_brand.name if current_brand else "미지정 브랜드"
                data.append([Paragraph(f"■ {brand_name}", bold_style), "", ""])
                table_style.add('SPAN', (0, row_idx), (-1, row_idx))
                table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f1f5f9'))
                table_style.add('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 6)
                table_style.add('TOPPADDING', (0, row_idx), (-1, row_idx), 6)
                row_idx += 1
            
            invs = Inventory.objects.filter(tool=tool, status='재고').order_by('date')
            count = invs.count()
            total_qty += count
            
            serials = [inv.serial for inv in invs if inv.serial]
            no_serial_count = invs.filter(Q(serial__isnull=True) | Q(serial__exact='')).count()
            parts = []
            if serials:
                parts.append(", ".join(serials))
            if no_serial_count > 0:
                parts.append(f"S/N 없음: {no_serial_count}개")
            
            serial_text = " / ".join(parts) if parts else "-"

            data.append([
                Paragraph(tool.model_name, normal_style),
                Paragraph(str(count), center_style),
                Paragraph(serial_text, normal_style)
            ])
            table_style.add('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 6)
            table_style.add('TOPPADDING', (0, row_idx), (-1, row_idx), 6)
            row_idx += 1

        data.append([
            Paragraph("<b>총합계 (Total)</b>", ParagraphStyle('T1', parent=bold_style, alignment=1)),
            Paragraph(f"<b>{total_qty}</b>", ParagraphStyle('T2', parent=bold_style, alignment=1)),
            ""
        ])
        table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f8fafc'))
        table_style.add('LINEABOVE', (0, row_idx), (-1, row_idx), 1, colors.HexColor('#94a3b8'))
        table_style.add('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 8)
        table_style.add('TOPPADDING', (0, row_idx), (-1, row_idx), 8)

        colWidths = [150, 60, 325]
        t = Table(data, colWidths=colWidths, repeatRows=1)
        t.setStyle(table_style)
        
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"재고현황_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        return FileResponse(buffer, as_attachment=True, filename=filename)

    def dashboard_history_pdf(self, request):
        """대시보드: 기간별 입출고 이력 PDF 다운로드"""
        import io, os
        from datetime import date, datetime
        from django.http import FileResponse
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from .models import Inventory
        from django.utils import timezone

        # 쿼리 파라미터에서 시작일/종료일 수신
        start_str = request.GET.get('start', '')
        end_str = request.GET.get('end', '')

        today = date.today()
        if start_str:
            try:
                start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            except ValueError:
                start_date = today.replace(day=1)
        else:
            start_date = today.replace(day=1)

        if end_str:
            try:
                end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
            except ValueError:
                end_date = today
        else:
            end_date = today

        # 해당 기간 내 입고 또는 출고된 모든 데이터
        qs_inbound = Inventory.objects.filter(
            date__gte=start_date, date__lte=end_date
        ).select_related('tool', 'tool__brand', 'supplier').order_by('date', 'tool__model_name')

        qs_outbound = Inventory.objects.filter(
            status='출고',
            release_date__gte=start_date, release_date__lte=end_date
        ).select_related('tool', 'tool__brand', 'release_company').order_by('release_date', 'tool__model_name')

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=40)
        elements = []

        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'
        try:
            font_path = "c:\\Windows\\Fonts\\malgun.ttf"
            font_path_bd = "c:\\Windows\\Fonts\\malgunbd.ttf"
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('Malgun', font_path))
                font_name = 'Malgun'
                font_name_bold = 'Malgun'
            if os.path.exists(font_path_bd):
                pdfmetrics.registerFont(TTFont('MalgunBold', font_path_bd))
                font_name_bold = 'MalgunBold'
        except Exception:
            pass

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Normal'],
            fontName=font_name_bold, fontSize=18, alignment=1, spaceAfter=15
        )
        date_style = ParagraphStyle(
            'CustomDate', parent=styles['Normal'],
            fontName=font_name, fontSize=10, textColor=colors.gray, alignment=2, spaceAfter=5
        )
        period_style = ParagraphStyle(
            'CustomPeriod', parent=styles['Normal'],
            fontName=font_name, fontSize=11, alignment=1, spaceAfter=25
        )
        normal_style = ParagraphStyle(
            'CustomNormal', parent=styles['Normal'],
            fontName=font_name, fontSize=9, leading=14
        )
        bold_style = ParagraphStyle(
            'CustomBold', parent=styles['Normal'],
            fontName=font_name_bold, fontSize=10, leading=14
        )
        center_style = ParagraphStyle(
            'CustomCenter', parent=normal_style, alignment=1
        )
        
        # Subtitles (Green for inbound, Blue for outbound)
        inbound_header_style = ParagraphStyle(
            'InboundHeader', parent=styles['Normal'],
            fontName=font_name_bold, fontSize=13, textColor=colors.HexColor('#059669'), spaceAfter=10
        )
        outbound_header_style = ParagraphStyle(
            'OutboundHeader', parent=styles['Normal'],
            fontName=font_name_bold, fontSize=13, textColor=colors.HexColor('#2563eb'), spaceAfter=10
        )

        period_label = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"

        elements.append(Paragraph("입출고 이력 보고서", title_style))
        elements.append(Paragraph(f"출력일시: {timezone.now().strftime('%Y-%m-%d %H:%M')}", date_style))
        elements.append(Paragraph(f"조회 기간: {period_label}", period_style))

        # Common table style
        base_table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#94a3b8')),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
        ])
        
        colWidths = [65, 150, 120, 200]

        # ─── 입고 내역 섹션 ───
        elements.append(Paragraph(f"■ 입고 내역 ({qs_inbound.count()}건)", inbound_header_style))
        
        if qs_inbound.exists():
            inbound_data = [[
                Paragraph("<b>입고일</b>", ParagraphStyle('H1', parent=center_style, fontName=font_name_bold)),
                Paragraph("<b>품목명 (Model)</b>", ParagraphStyle('H2', parent=normal_style, fontName=font_name_bold)),
                Paragraph("<b>입고처 (Supplier)</b>", ParagraphStyle('H3', parent=normal_style, fontName=font_name_bold)),
                Paragraph("<b>시리얼 내역 (S/N)</b>", ParagraphStyle('H4', parent=normal_style, fontName=font_name_bold))
            ]]
            for inv in qs_inbound:
                date_str = inv.date.strftime('%Y-%m-%d') if inv.date else "-"
                tool_name = str(inv.tool) if inv.tool else "-"
                supplier_name = inv.supplier.name if inv.supplier else "-"
                serial = inv.serial if inv.serial else "-"
                
                inbound_data.append([
                    Paragraph(date_str, center_style),
                    Paragraph(tool_name, normal_style),
                    Paragraph(supplier_name, normal_style),
                    Paragraph(serial, normal_style)
                ])
                
            t_in = Table(inbound_data, colWidths=colWidths, repeatRows=1)
            t_in.setStyle(base_table_style)
            elements.append(t_in)
        else:
            elements.append(Paragraph("해당 기간 내 입고 내역이 없습니다.", normal_style))
            
        elements.append(Spacer(1, 30))

        # ─── 출고 내역 섹션 ───
        elements.append(Paragraph(f"■ 출고 내역 ({qs_outbound.count()}건)", outbound_header_style))
        
        if qs_outbound.exists():
            outbound_data = [[
                Paragraph("<b>출고일</b>", ParagraphStyle('H1', parent=center_style, fontName=font_name_bold)),
                Paragraph("<b>품목명 (Model)</b>", ParagraphStyle('H2', parent=normal_style, fontName=font_name_bold)),
                Paragraph("<b>출고처 (Company)</b>", ParagraphStyle('H3', parent=normal_style, fontName=font_name_bold)),
                Paragraph("<b>시리얼 내역 (S/N)</b>", ParagraphStyle('H4', parent=normal_style, fontName=font_name_bold))
            ]]
            for inv in qs_outbound:
                date_str = inv.release_date.strftime('%Y-%m-%d') if inv.release_date else "-"
                tool_name = str(inv.tool) if inv.tool else "-"
                company_name = inv.release_company.name if inv.release_company else "-"
                serial = inv.serial if inv.serial else "-"
                
                outbound_data.append([
                    Paragraph(date_str, center_style),
                    Paragraph(tool_name, normal_style),
                    Paragraph(company_name, normal_style),
                    Paragraph(serial, normal_style)
                ])
                
            t_out = Table(outbound_data, colWidths=colWidths, repeatRows=1)
            t_out.setStyle(base_table_style)
            elements.append(t_out)
        else:
            elements.append(Paragraph("해당 기간 내 출고 내역이 없습니다.", normal_style))

        doc.build(elements)
        buffer.seek(0)
        filename = f"기간별_입출고이력_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}_{timezone.now().strftime('%H%M')}.pdf"
        return FileResponse(buffer, as_attachment=True, filename=filename)

tool_admin_site = ToolInventoryAdminSite(name='tool_admin')

# ── Register AS_APP models for autocomplete functionality in Tool Inventory ──
# NOTE: Company/OutsourceCompany are registered here ONLY for autocomplete support.
# changelist_view/change_view are redirected to prevent NoReverseMatch errors
# caused by company_tabs.html referencing URLs not registered in tool_admin_site.
@admin.register(OutsourceCompany, site=tool_admin_site)
class ToolOutsourceCompanyAdmin(ModelAdmin):
    search_fields = ["name"]
    def has_module_permission(self, request): return False
    def changelist_view(self, request, extra_context=None):
        return HttpResponseRedirect(reverse('tool_admin:index'))
    def change_view(self, request, object_id, form_url='', extra_context=None):
        return HttpResponseRedirect(reverse('tool_admin:index'))

@admin.register(Company, site=tool_admin_site)
class ToolCompanyAdmin(ModelAdmin):
    search_fields = ["name"]
    def has_module_permission(self, request): return False
    def changelist_view(self, request, extra_context=None):
        return HttpResponseRedirect(reverse('tool_admin:index'))
    def change_view(self, request, object_id, form_url='', extra_context=None):
        return HttpResponseRedirect(reverse('tool_admin:index'))

@admin.register(Brand, site=tool_admin_site)
class ToolBrandAdmin(ModelAdmin):
    search_fields = ["name"]
    def has_module_permission(self, request): return False

@admin.register(Tool, site=tool_admin_site)
class ToolToolAdmin(ModelAdmin):
    search_fields = ["model_name", "brand__name"]
    autocomplete_fields = ["brand"]
    def has_module_permission(self, request): return False

@admin.register(Inventory, site=tool_admin_site)
class InventoryAdmin(ModelAdmin):
    list_display = (
        'display_status', 'date', 'supplier_text', 'release_date', 'release_company', 
        'tool_text', 'serial', 'usage_process', 'display_edit_button'
    )
    list_display_links = None
    actions = ['cancel_outbound', 'export_selected_to_pdf', 'export_selected_to_excel']
    list_filter = ('status', 'supplier', 'release_company')
    search_fields = ('tool__model_name', 'tool__brand__name', 'serial', 'supplier__name', 'release_company__name', 'usage_process')
    list_per_page = 50
    autocomplete_fields = ('tool', 'supplier', 'release_company')

    def get_ordering(self, request):
        """필터 컨텍스트에 따른 동적 정렬:
        - 출고 상태 필터 또는 출고처 필터 선택 시 → 출고일자 최신순 우선
        - 입고처 필터 선택 시 → 입고일자 최신순 우선
        - 기본 → 입고일자 최신순 + 품목명 오름차순
        """
        status_filter = request.GET.get('status__exact', '')
        supplier_filter = request.GET.get('supplier__id__exact', '')
        release_company_filter = request.GET.get('release_company__id__exact', '')

        if release_company_filter or status_filter == '출고':
            # 출고처 필터 또는 출고 상태 필터 → 출고일자 최신순 우선
            return ['-release_date', 'tool__model_name']
        elif supplier_filter:
            # 입고처 필터 → 입고일자 최신순 우선
            return ['-date', 'tool__model_name']

        # 기본 정렬: 입고일자 최신순 + 품목명 오름차순
        return ['-date', 'tool__model_name']

    class Media:
        css = {"all": ("as_app/css/row_colors.css",)}
        js = ("as_app/js/inventory_change.js",)

    @unfold_action(description="↩️ 선택한 장비 출고 취소 (재고로 원복)")
    def cancel_outbound(self, request, queryset):
        qs = queryset.filter(status='출고')
        count = qs.count()
        if count == 0:
            from django.contrib import messages
            messages.warning(request, "선택한 항목 중 '출고' 상태인 장비가 없습니다.")
            return

        for obj in qs:
            obj.status = '재고'
            obj.release_date = None
            obj.release_company = None
            obj.save(update_fields=['status', 'release_date', 'release_company'])
            
        from django.contrib import messages
        messages.success(request, f"총 {count}개의 장비가 성공적으로 재고로 원복되었습니다.")

    @unfold_action(description="📄 선택한 내역 PDF 다운로드", url_path="export-selected-pdf", attrs={"style": "background-color: #9333ea; color: white; border: none;"})
    def export_selected_to_pdf(self, request, queryset):
        import io, os
        from django.http import FileResponse
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from django.utils import timezone
        from itertools import groupby
        
        # 1. 문서 메타데이터 설정
        queryset = queryset.order_by('tool__brand__name', 'tool__model_name', 'date')
        
        outbound_companies = list(set([obj.release_company.name for obj in queryset if obj.release_company and obj.status == '출고']))
        is_all_outbound = all(obj.status == '출고' for obj in queryset)
        
        if is_all_outbound and len(outbound_companies) == 1:
            title = f"({outbound_companies[0]}) 출고리스트"
        elif is_all_outbound and len(outbound_companies) > 1:
            title = "(다중업체) 출고리스트"
        else:
            title = "입출고 이력 내역"

        dates_set = set()
        for obj in queryset:
            if is_all_outbound:
                if obj.release_date:
                    dates_set.add(obj.release_date.strftime('%Y-%m-%d'))
            else:
                if obj.date:
                    dates_set.add(obj.date.strftime('%Y-%m-%d'))
        
        dates_list = sorted(list(dates_set))
        if dates_list:
            date_range = ", ".join(dates_list)
        else:
            date_range = ""

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=40)
        elements = []
        
        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'
        try:
            font_path = "c:\\Windows\\Fonts\\malgun.ttf"
            font_path_bd = "c:\\Windows\\Fonts\\malgunbd.ttf"
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('Malgun', font_path))
                font_name = 'Malgun'
                font_name_bold = 'Malgun'
            if os.path.exists(font_path_bd):
                pdfmetrics.registerFont(TTFont('MalgunBold', font_path_bd))
                font_name_bold = 'MalgunBold'
        except Exception:
            pass

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Normal'],
            fontName=font_name_bold, fontSize=18, alignment=1, spaceAfter=5
        )
        date_range_style = ParagraphStyle(
            'CustomDateRange', parent=styles['Normal'],
            fontName=font_name, fontSize=11, alignment=1, spaceAfter=20
        )
        date_style = ParagraphStyle(
            'CustomDate', parent=styles['Normal'],
            fontName=font_name, fontSize=10, textColor=colors.gray, alignment=2, spaceAfter=15
        )
        normal_style = ParagraphStyle(
            'CustomNormal', parent=styles['Normal'],
            fontName=font_name, fontSize=9, leading=14
        )
        bold_style = ParagraphStyle(
            'CustomBold', parent=styles['Normal'],
            fontName=font_name_bold, fontSize=10, leading=14
        )
        center_style = ParagraphStyle(
            'CustomCenter', parent=normal_style, alignment=1
        )
        
        # 문서 헤더 구성
        elements.append(Paragraph(title, title_style))
        if date_range:
            elements.append(Paragraph(f"({date_range})", date_range_style))
        else:
            elements.append(Spacer(1, 15))
            
        elements.append(Paragraph(f"출력일시: {timezone.now().strftime('%Y-%m-%d %H:%M')}", date_style))
        
        # 테이블 스타일 구성
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#94a3b8')),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
        ])

        # 테이블 헤더
        data = [[
            Paragraph("<b>품목명 (Model)</b>", ParagraphStyle('H1', parent=normal_style, fontName=font_name_bold)),
            Paragraph("<b>수량 (Qty)</b>", ParagraphStyle('H2', parent=center_style, fontName=font_name_bold)),
            Paragraph("<b>시리얼 내역 (S/N)</b>", ParagraphStyle('H3', parent=normal_style, fontName=font_name_bold))
        ]]

        def get_tool_name(obj):
            return str(obj.tool) if obj.tool else "미지정 품목"

        items = list(queryset)
        items.sort(key=get_tool_name)
        grouped_items = groupby(items, key=get_tool_name)
        
        row_idx = 1
        total_qty = 0
        current_brand = None
        
        for tool_name, group in grouped_items:
            group_list = list(group)
            quantity = len(group_list)
            total_qty += quantity
            
            first_obj = group_list[0]
            brand_name = first_obj.tool.brand.name if first_obj.tool and first_obj.tool.brand else "미지정 브랜드"
            model_name = first_obj.tool.model_name if first_obj.tool else "미지정 품목"
            
            if brand_name != current_brand:
                current_brand = brand_name
                data.append([Paragraph(f"■ {current_brand}", bold_style), "", ""])
                table_style.add('SPAN', (0, row_idx), (-1, row_idx))
                table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f1f5f9'))
                table_style.add('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 6)
                table_style.add('TOPPADDING', (0, row_idx), (-1, row_idx), 6)
                row_idx += 1
            
            # 시리얼 번호 수집
            serials = []
            no_serial_count = 0
            for obj in group_list:
                if obj.serial:
                    serials.append(obj.serial)
                else:
                    no_serial_count += 1
            
            parts = []
            if serials:
                parts.append(", ".join(serials))
            if no_serial_count > 0:
                parts.append(f"S/N 없음: {no_serial_count}개")
            
            serial_text = " / ".join(parts) if parts else "-"
            
            data.append([
                Paragraph(model_name, normal_style),
                Paragraph(str(quantity), center_style),
                Paragraph(serial_text, normal_style)
            ])
            table_style.add('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 6)
            table_style.add('TOPPADDING', (0, row_idx), (-1, row_idx), 6)
            row_idx += 1

        # 총합
        data.append([
            Paragraph("<b>총합계 (Total)</b>", ParagraphStyle('T1', parent=bold_style, alignment=1)),
            Paragraph(f"<b>{total_qty}</b>", ParagraphStyle('T2', parent=bold_style, alignment=1)),
            ""
        ])
        table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f8fafc'))
        table_style.add('LINEABOVE', (0, row_idx), (-1, row_idx), 1, colors.HexColor('#94a3b8'))
        table_style.add('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 8)
        table_style.add('TOPPADDING', (0, row_idx), (-1, row_idx), 8)
        
        colWidths = [150, 60, 325]
        t = Table(data, colWidths=colWidths, repeatRows=1)
        t.setStyle(table_style)
        
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        
        safe_title = title.replace(' ', '_').replace('()', '')
        filename = f"{safe_title}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        response = FileResponse(buffer, as_attachment=True, filename=filename)
        return response

    @unfold_action(description="📊 선택한 내역 엑셀 다운로드", url_path="export-selected-excel", attrs={"style": "background-color: #10b981; color: white; border: none;"})
    def export_selected_to_excel(self, request, queryset):
        import io
        import urllib.parse
        import openpyxl
        from django.http import HttpResponse
        from django.utils import timezone
        from openpyxl.styles import Font, Alignment
        
        # 품명 기준으로 정렬 (브랜드 -> 모델명 우선)
        queryset = queryset.order_by('tool__brand__name', 'tool__model_name', 'date')
        
        # 출고 업체명 확인하여 파일명 구성
        outbound_companies = list(set([obj.release_company.name for obj in queryset if obj.release_company and obj.status == '출고']))
        is_all_outbound = all(obj.status == '출고' for obj in queryset)
        
        if is_all_outbound and len(outbound_companies) == 1:
            title = f"({outbound_companies[0]}) 출고리스트"
        elif is_all_outbound and len(outbound_companies) > 1:
            title = "(다중업체) 출고리스트"
        else:
            title = "입출고이력내역"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "입출고이력"
        
        headers = ["브랜드", "모델명", "시리얼번호", "상태", "처리일자", "거래처"]
        ws.append(headers)
        
        # 헤더 스타일
        header_font = Font(bold=True)
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for obj in queryset:
            brand_name = obj.tool.brand.name if obj.tool and obj.tool.brand else "-"
            model_name = obj.tool.model_name if obj.tool else "미지정 품목"
            serial_text = obj.serial if obj.serial else "S/N 없음"
            status_text = obj.get_status_display()
            
            if obj.status == '출고':
                date_str = str(obj.release_date) if obj.release_date else "-"
                company_text = obj.release_company.name if obj.release_company else "-"
            else:
                date_str = str(obj.date) if obj.date else "-"
                company_text = obj.supplier.name if obj.supplier else "-"
                
            ws.append([brand_name, model_name, serial_text, status_text, date_str, company_text])
            
        # 열 너비 설정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        safe_title = title.replace(' ', '_').replace('()', '')
        filename = f"{safe_title}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
        return response

    def has_add_permission(self, request):
        return False

    def render_change_form(self, request, context, add=False, change=False, form_url="", obj=None):
        context["show_save_and_continue"] = False
        return super().render_change_form(request, context, add, change, form_url, obj)

    def get_fields(self, request, obj=None):
        if obj:
            return ('id', 'date', 'supplier_text', 'tool_text', 'serial', 'release_date', 'release_company', 'usage_process', 'status')
        return ('date', 'supplier', 'tool', 'serial', 'usage_process', 'status')

    def get_readonly_fields(self, request, obj=None):
        if obj:
            # 수정 모드: 'serial', 'release_company', 'release_date', 'usage_process' 수정 가능
            # 입고배치(batch)는 fields에서 제외하여 보이지 않게 함
            # 품목(tool) 및 입고처(supplier)는 단순 텍스트로 표시되어 링크 이동 방지
            return ['id', 'date', 'supplier_text', 'tool_text', 'status']
        return super().get_readonly_fields(request, obj)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if 'release_company' in form.base_fields:
            widget = form.base_fields['release_company'].widget
            widget.can_add_related = False
            widget.can_change_related = False
            widget.can_delete_related = False
            widget.can_view_related = False
        return form

    @display(description="입고처")
    def supplier_text(self, obj):
        return str(obj.supplier) if obj.supplier else "-"

    @display(description="품목명")
    def tool_text(self, obj):
        return str(obj.tool) if obj.tool else "-"

    @display(description="상태")
    def display_status(self, obj):
        if obj.status == "재고":
            return obj.get_status_display()
        return format_html(
            '<span class="status-marker" data-status="repaired">{}</span>',
            obj.get_status_display(),
        )

    @display(description="수정")
    def display_edit_button(self, obj):
        url = reverse("tool_admin:tool_inventory_inventory_change", args=[obj.pk])
        return format_html(
            '<a href="{}" style="'
            'display:inline-flex; align-items:center; gap:4px; '
            'padding:6px 12px; border-radius:6px; font-size:0.8rem; '
            'font-weight:600; text-decoration:none; '
            'background:#8b5cf6; color:#fff; '
            'transition:all .15s ease;'
            '" '
            'onmouseover="this.style.background=\'#7c3aed\'" '
            'onmouseout="this.style.background=\'#8b5cf6\'">'
            '📝 수정</a>',
            url
        )

class InventoryInline(TabularInline):
    model = Inventory
    fk_name = "batch"
    form = InventoryForm
    fields = ["brand", "tool", "no_serial", "quantity", "serial"]
    verbose_name = "입고등록 티켓"
    verbose_name_plural = "입고등록 티켓"
    extra = 0
    min_num = 1

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        for field in formset.form.base_fields.values():
            if hasattr(field, 'widget'):
                field.widget.can_add_related = False
                field.widget.can_change_related = False
                field.widget.can_delete_related = False
                field.widget.can_view_related = False
                if hasattr(field.widget, 'attrs'):
                    field.widget.attrs['autocomplete'] = 'off'
        if 'serial' in formset.form.base_fields:
            formset.form.base_fields['serial'].label = "시리얼 번호 (쉼표로 구분 시 다건 등록)"
        return formset

@admin.register(InventoryBatch, site=tool_admin_site)
class InventoryBatchAdmin(ModelAdmin):
    list_display = ["inbound_date", "supplier", "created_at"]
    list_filter = ["inbound_date", "supplier"]
    search_fields = ["supplier__name"]
    autocomplete_fields = ["supplier"]
    date_hierarchy = "inbound_date"
    list_per_page = 20
    inlines = [InventoryInline]

    class Media:
        css = {"all": ("as_app/css/inline_fix.css", "as_app/css/hide_fab.css")}
        js = ("as_app/js/inbound_form.js",)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if 'supplier' in form.base_fields:
            widget = form.base_fields['supplier'].widget
            widget.can_add_related = False
            widget.can_change_related = False
            widget.can_delete_related = False
            widget.can_view_related = False
        return form

    def render_change_form(self, request, context, add=False, change=False, form_url="", obj=None):
        context["show_save_and_add_another"] = False
        context["show_save_and_continue"] = False
        return super().render_change_form(request, context, add, change, form_url, obj)

    def response_add(self, request, obj, post_url_continue=None):
        if getattr(request, "_formset_validation_failed", False):
            if obj.pk and obj.inventories.count() == 0:
                obj.delete()
            return HttpResponseRedirect(request.path)
        from django.contrib import messages
        messages.success(request, f"입고 등록이 완료되었습니다. ({obj})")
        return HttpResponseRedirect(request.path)

    def response_change(self, request, obj):
        if getattr(request, "_formset_validation_failed", False):
            return HttpResponseRedirect(request.path)
        from django.contrib import messages
        messages.success(request, f"입고 정보가 수정되었습니다. ({obj})")
        return HttpResponseRedirect(request.path)

    def save_formset(self, request, form, formset, change):
        instances = formset.save(commit=False)
        for obj in formset.deleted_objects:
            obj.delete()
        batch = form.instance
        
        expanded = []
        for inline_form in formset.forms:
            if inline_form.instance in formset.deleted_objects or getattr(inline_form, 'cleaned_data', None) is None:
                continue
            if not inline_form.has_changed() and not inline_form.instance.pk:
                continue
                
            instance = inline_form.instance
            no_serial = inline_form.cleaned_data.get('no_serial', False)
            quantity = inline_form.cleaned_data.get('quantity', 1) or 1
            
            if no_serial:
                for _ in range(quantity):
                    if _ == 0:
                        instance.serial = None
                        expanded.append(instance)
                    else:
                        new_inv = Inventory(
                            batch=batch,
                            tool=instance.tool,
                            serial=None,
                        )
                        expanded.append(new_inv)
            elif instance.tool_id and instance.serial:
                serials = [s.strip() for s in instance.serial.split(",") if s.strip()]
                if len(serials) > 1:
                    instance.serial = serials[0]
                    expanded.append(instance)
                    for sn in serials[1:]:
                        new_inv = Inventory(
                            batch=batch,
                            tool=instance.tool,
                            serial=sn,
                        )
                        expanded.append(new_inv)
                else:
                    expanded.append(instance)
            elif instance.tool_id:
                expanded.append(instance)

        saved_count = 0
        for instance in expanded:
            instance.date = batch.inbound_date
            instance.supplier = batch.supplier
            instance.save()
            saved_count += 1
        formset.save_m2m()
        
        original_count = len(instances)
        if saved_count > original_count:
            from django.contrib import messages
            messages.info(request, f"시리얼번호 분리로 총 {saved_count}건의 툴장비가 등록되었습니다.")

from .models import Inventory, InventoryBatch, InboundInventory, OutboundInventory, OutboundBatch, OutboundTicket
from .forms import InventoryForm, OutboundTicketForm

class OutboundTicketInline(TabularInline):
    model = OutboundTicket
    fk_name = "batch"
    form = OutboundTicketForm
    fields = ["brand", "tool", "current_stock", "quantity", "inventories", "usage_process"]
    verbose_name = "출고등록 티켓"
    verbose_name_plural = "출고등록 티켓"
    extra = 0
    min_num = 1

@admin.register(OutboundBatch, site=tool_admin_site)
class OutboundBatchAdmin(ModelAdmin):
    list_display = ["release_date", "release_company", "created_at"]
    list_filter = ["release_date", "release_company"]
    search_fields = ["release_company__name"]
    autocomplete_fields = ["release_company"]
    list_per_page = 20
    inlines = [OutboundTicketInline]

    class Media:
        css = {"all": ("as_app/css/inline_fix.css", "as_app/css/hide_fab.css", "as_app/css/outbound_form.css")}
        js = ("as_app/js/outbound_form.js",)

    def get_model_perms(self, request):
        return {}

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        
    def save_formset(self, request, form, formset, change):
        super().save_formset(request, form, formset, change)
        if formset.model == OutboundTicket:
            for inline_form in formset.forms:
                instance = inline_form.instance
                
                # 삭제 대상이거나 빈 폼(cleaned_data가 없는 경우)은 건너뜀
                if instance in formset.deleted_objects or not getattr(inline_form, 'cleaned_data', None):
                    continue
                
                selected_inventories = list(instance.inventories.all())
                
                # 시리얼 번호가 골라지지 않았을 때는 선입선출 자동 배정
                if not selected_inventories:
                    qty = instance.quantity or 1
                    auto_inventories = Inventory.objects.filter(
                        tool=instance.tool, 
                        status='재고'
                    ).order_by('date')[:qty]
                    
                    for inv in auto_inventories:
                        inv.release_date = instance.batch.release_date
                        inv.release_company = instance.batch.release_company
                        inv.status = '출고'
                        inv.usage_process = instance.usage_process
                        inv.save()
                        instance.inventories.add(inv)
                else:
                    # 유저가 직접 다중 선택한 경우
                    for inv in selected_inventories:
                        inv.release_date = instance.batch.release_date
                        inv.release_company = instance.batch.release_company
                        inv.status = '출고'
                        inv.usage_process = instance.usage_process
                        inv.save()
                        
                instance.quantity = instance.inventories.count()
                instance.save()

@admin.register(OutboundInventory, site=tool_admin_site)
class OutboundInventoryAdmin(ModelAdmin):
    list_display = (
        'display_status', 'date', 'supplier', 'tool', 'serial'
    )
    list_display_links = None
    list_filter = ["supplier", "tool__brand", "tool"]
    search_fields = ["serial", "supplier__name", "tool__model_name", "tool__brand__name"]
    list_per_page = 30
    
    def changelist_view(self, request, extra_context=None):
        return HttpResponseRedirect(reverse("tool_admin:tool_inventory_outboundbatch_add"))
    list_display = (
        'display_status', 'date', 'supplier', 'tool', 'serial'
    )
    list_display_links = None
    list_filter = ["supplier", "tool__brand", "tool"]
    search_fields = ["serial", "supplier__name", "tool__model_name", "tool__brand__name"]
    list_per_page = 30

    class Media:
        css = {"all": ("as_app/css/row_colors.css",)}

    @display(description="상태")
    def display_status(self, obj):
        if obj.status == "재고":
            return obj.get_status_display()
        return format_html(
            '<span class="status-marker" data-status="repaired">{}</span>',
            obj.get_status_display(),
        )

    def get_queryset(self, request):
        # '재고' 상태인 데이터만 표시
        return super().get_queryset(request).filter(status='재고')

    def add_view(self, request, form_url='', extra_context=None):
        # "추가" 버튼 클릭 시 OutboundBatch 생성 화면으로 리다이렉트
        return HttpResponseRedirect(reverse("tool_admin:tool_inventory_outboundbatch_add"))

from .models import ToolStockSummary

@admin.register(ToolStockSummary, site=tool_admin_site)
class ToolStockSummaryAdmin(ModelAdmin):
    list_display = ('brand', 'model_name', 'stock_count', 'serial_list')
    search_fields = ('brand__name', 'model_name')
    list_filter = ('brand',)
    list_per_page = 50
    actions = ["export_stock_pdf", "export_stock_excel"]

    def has_add_permission(self, request):
        return False
        
    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        # 재고가 있는 장비만 표시
        return super().get_queryset(request).filter(inventory__status='재고').distinct()

    @display(description="현재 재고 수량")
    def stock_count(self, obj):
        count = Inventory.objects.filter(tool=obj, status='재고').count()
        return format_html(
            '<span style="font-weight: bold; color: {};">{}개</span>',
            '#10b981' if count > 0 else '#ef4444', 
            count
        )

    @display(description="보유 시리얼 목록")
    def serial_list(self, obj):
        from django.db.models import Q
        invs = Inventory.objects.filter(tool=obj, status='재고').order_by('date')
        serials = [inv.serial for inv in invs if inv.serial]
        no_serial_count = invs.filter(Q(serial__isnull=True) | Q(serial__exact='')).count()
        
        parts = []
        if serials:
            parts.append(", ".join(serials))
        if no_serial_count > 0:
            parts.append(f"(S/N 없음: {no_serial_count}개)")
            
        return " / ".join(parts) if parts else "-"

    @unfold_action(description="재고 내역 출력 (PDF)", attrs={"style": "background-color: #9333ea; color: white; border: none;"})
    def export_stock_pdf(self, request, queryset):
        import io, os
        from django.http import FileResponse
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from django.db.models import Q
        from .models import Inventory
        from django.utils import timezone
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=40)
        elements = []
        
        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'
        try:
            font_path = "c:\\Windows\\Fonts\\malgun.ttf"
            font_path_bd = "c:\\Windows\\Fonts\\malgunbd.ttf"
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('Malgun', font_path))
                font_name = 'Malgun'
                font_name_bold = 'Malgun'
            if os.path.exists(font_path_bd):
                pdfmetrics.registerFont(TTFont('MalgunBold', font_path_bd))
                font_name_bold = 'MalgunBold'
        except Exception:
            pass
            
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Normal'],
            fontName=font_name_bold, fontSize=18, alignment=1, spaceAfter=20
        )
        date_style = ParagraphStyle(
            'CustomDate', parent=styles['Normal'],
            fontName=font_name, fontSize=10, textColor=colors.gray, alignment=2, spaceAfter=15
        )
        normal_style = ParagraphStyle(
            'CustomNormal', parent=styles['Normal'],
            fontName=font_name, fontSize=9, leading=14
        )
        bold_style = ParagraphStyle(
            'CustomBold', parent=styles['Normal'],
            fontName=font_name_bold, fontSize=10, leading=14
        )
        center_style = ParagraphStyle(
            'CustomCenter', parent=normal_style, alignment=1
        )

        elements.append(Paragraph("선택 장비 재고 내역", title_style))
        elements.append(Paragraph(f"출력일시: {timezone.now().strftime('%Y-%m-%d %H:%M')}", date_style))

        queryset = queryset.select_related('brand').order_by('brand__name', 'model_name')
        
        data = []
        data.append([
            Paragraph("<b>품목명 (Model)</b>", ParagraphStyle('H1', parent=normal_style, fontName=font_name_bold)),
            Paragraph("<b>수량 (Qty)</b>", ParagraphStyle('H2', parent=center_style, fontName=font_name_bold)),
            Paragraph("<b>시리얼 내역 (S/N)</b>", ParagraphStyle('H3', parent=normal_style, fontName=font_name_bold))
        ])

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#94a3b8')),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
        ])

        current_brand = None
        total_qty = 0
        row_idx = 1
        
        for tool in queryset:
            if tool.brand != current_brand:
                current_brand = tool.brand
                brand_name = current_brand.name if current_brand else "미지정 브랜드"
                data.append([Paragraph(f"■ {brand_name}", bold_style), "", ""])
                table_style.add('SPAN', (0, row_idx), (-1, row_idx))
                table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f1f5f9'))
                table_style.add('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 6)
                table_style.add('TOPPADDING', (0, row_idx), (-1, row_idx), 6)
                row_idx += 1
            
            invs = Inventory.objects.filter(tool=tool, status='재고').order_by('date')
            count = invs.count()
            total_qty += count
            
            serials = [inv.serial for inv in invs if inv.serial]
            no_serial_count = invs.filter(Q(serial__isnull=True) | Q(serial__exact='')).count()
            
            parts = []
            if serials:
                parts.append(", ".join(serials))
            if no_serial_count > 0:
                parts.append(f"S/N 없음: {no_serial_count}개")
            
            serial_text = " / ".join(parts) if parts else "-"

            data.append([
                Paragraph(tool.model_name, normal_style),
                Paragraph(str(count), center_style),
                Paragraph(serial_text, normal_style)
            ])
            table_style.add('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 6)
            table_style.add('TOPPADDING', (0, row_idx), (-1, row_idx), 6)
            row_idx += 1

        data.append([
            Paragraph("<b>총합계 (Total)</b>", ParagraphStyle('T1', parent=bold_style, alignment=1)),
            Paragraph(f"<b>{total_qty}</b>", ParagraphStyle('T2', parent=bold_style, alignment=1)),
            ""
        ])
        table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f8fafc'))
        table_style.add('LINEABOVE', (0, row_idx), (-1, row_idx), 1, colors.HexColor('#94a3b8'))
        table_style.add('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 8)
        table_style.add('TOPPADDING', (0, row_idx), (-1, row_idx), 8)

        colWidths = [150, 60, 325]
        t = Table(data, colWidths=colWidths, repeatRows=1)
        t.setStyle(table_style)
        
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"선택_재고내역_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        return FileResponse(buffer, as_attachment=True, filename=filename)

    @unfold_action(description="재고 내역 출력 (Excel)", attrs={"style": "background-color: #10b981; color: white; border: none;"})
    def export_stock_excel(self, request, queryset):
        import io
        import openpyxl
        from django.http import HttpResponse
        from django.utils import timezone
        from openpyxl.styles import Font, Alignment
        from django.db.models import Q
        
        queryset = queryset.select_related('brand').order_by('brand__name', 'model_name')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "재고현황"
        
        headers = ["브랜드", "모델명", "수량", "재고 내역(시리얼)"]
        ws.append(headers)
        
        header_font = Font(bold=True)
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for tool in queryset:
            brand_name = tool.brand.name if tool.brand else "미지정"
            model_name = tool.model_name
            
            invs = Inventory.objects.filter(tool=tool, status='재고').order_by('date')
            count = invs.count()
            
            serials = [inv.serial for inv in invs if inv.serial]
            no_serial_count = invs.filter(Q(serial__isnull=True) | Q(serial__exact='')).count()
            
            parts = []
            if serials:
                parts.append(", ".join(serials))
            if no_serial_count > 0:
                parts.append(f"S/N 없음: {no_serial_count}개")
            
            serial_text = " / ".join(parts) if parts else "재고 없음"
            
            ws.append([brand_name, model_name, count, serial_text])
            
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 50
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        import urllib.parse
        filename = f"선택_재고내역_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
        return response

# ── 투두리스트 관리 ──
from .models import TodoItem
from django.http import JsonResponse
import json

@admin.register(TodoItem, site=tool_admin_site)
class TodoItemAdmin(ModelAdmin):
    list_display = ["title", "is_done", "created_at"]
    list_filter = ["is_done"]
    search_fields = ["title"]

    def has_module_permission(self, request):
        """사이드바에 표시하지 않음 (대시보드에서만 관리)"""
        return False

    def get_urls(self):
        custom_urls = [
            path(
                "api/add/",
                self.admin_site.admin_view(self.api_add),
                name="todoitem_api_add",
            ),
            path(
                "api/toggle/<int:pk>/",
                self.admin_site.admin_view(self.api_toggle),
                name="todoitem_api_toggle",
            ),
            path(
                "api/delete/<int:pk>/",
                self.admin_site.admin_view(self.api_delete),
                name="todoitem_api_delete",
            ),
        ]
        return custom_urls + super().get_urls()

    def api_add(self, request):
        if request.method != "POST":
            return JsonResponse({"ok": False}, status=405)
        try:
            data = json.loads(request.body)
            title = data.get("title", "").strip()
            if not title:
                return JsonResponse({"ok": False, "error": "title required"}, status=400)
            todo = TodoItem.objects.create(title=title)
            return JsonResponse({"ok": True, "id": todo.id, "title": todo.title})
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)}, status=500)

    def api_toggle(self, request, pk):
        if request.method != "POST":
            return JsonResponse({"ok": False}, status=405)
        try:
            todo = TodoItem.objects.get(pk=pk)
            todo.is_done = not todo.is_done
            todo.save(update_fields=["is_done"])
            return JsonResponse({"ok": True, "is_done": todo.is_done})
        except TodoItem.DoesNotExist:
            return JsonResponse({"ok": False}, status=404)

    def api_delete(self, request, pk):
        if request.method != "POST":
            return JsonResponse({"ok": False}, status=405)
        try:
            TodoItem.objects.filter(pk=pk).delete()
            return JsonResponse({"ok": True})
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)}, status=500)
